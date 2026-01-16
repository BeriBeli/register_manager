"""
Excel Parser for Register Manager (Python/WASM Version)

This module parses Excel files containing register definitions and converts
them to IP-XACT compatible JSON format.
"""

import json
import re
from io import BytesIO
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook


def parse_bit_range(bit_str: str) -> tuple[int, int]:
    """
    Parse bit range string like '[7:0]', '7:0', '[3]', or '3'.
    Returns (msb, lsb) tuple.
    """
    bit_str = bit_str.strip()
    # Remove brackets if present
    bit_str = bit_str.strip('[]')
    
    if ':' in bit_str:
        parts = bit_str.split(':')
        msb = int(parts[0].strip())
        lsb = int(parts[1].strip())
        return (msb, lsb)
    else:
        # Single bit
        bit = int(bit_str.strip())
        return (bit, bit)


def parse_address(addr_str: str) -> int:
    """Parse address string (supports hex with 0x prefix or decimal)."""
    addr_str = addr_str.strip()
    if addr_str.startswith('0x') or addr_str.startswith('0X'):
        return int(addr_str, 16)
    return int(addr_str)


def parse_reset_value(reset_str: str) -> int:
    """Parse reset value (supports hex with 0x prefix or decimal)."""
    if not reset_str or reset_str.strip() == '':
        return 0
    reset_str = reset_str.strip()
    if reset_str.startswith('0x') or reset_str.startswith('0X'):
        return int(reset_str, 16)
    return int(reset_str)


def parse_version_sheet(ws) -> Dict[str, str]:
    """
    Parse the Version sheet.
    Expected format:
    Row 1: Header | Vendor | Library | Name | Version
    Row 2: Data   | ...    | ...     | ...  | ...
    """
    version_info = {
        'vendor': '',
        'library': '',
        'name': '',
        'version': ''
    }
    
    # Read row 2 (index 2 in openpyxl, 1-indexed)
    if ws.max_row >= 2:
        row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        if len(row) >= 5:
            version_info['vendor'] = str(row[1] or '')
            version_info['library'] = str(row[2] or '')
            version_info['name'] = str(row[3] or '')
            version_info['version'] = str(row[4] or '')
    
    return version_info


def parse_address_map_sheet(ws) -> List[Dict[str, Any]]:
    """
    Parse the Address Map sheet.
    Expected columns: Block Name | Offset | Range
    """
    blocks = []
    
    # Skip header row, start from row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        block_name = str(row[0]).strip()
        offset = parse_address(str(row[1])) if row[1] else 0
        range_val = parse_address(str(row[2])) if row[2] else 0
        
        blocks.append({
            'name': block_name,
            'baseAddress': offset,
            'range': range_val
        })
    
    return blocks


def parse_block_sheet(ws, block_name: str) -> List[Dict[str, Any]]:
    """
    Parse a block sheet containing register definitions.
    Expected columns: address | register_name | field_name | bit_offset | access | reset_value | description
    """
    registers = {}
    current_register = None
    current_address = None
    
    # Skip header row, start from row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Extract cell values
        addr_cell = row[0] if len(row) > 0 else None
        reg_name_cell = row[1] if len(row) > 1 else None
        field_name_cell = row[2] if len(row) > 2 else None
        bit_offset_cell = row[3] if len(row) > 3 else None
        access_cell = row[4] if len(row) > 4 else None
        reset_cell = row[5] if len(row) > 5 else None
        desc_cell = row[6] if len(row) > 6 else None
        
        # Skip completely empty rows
        if not any([addr_cell, reg_name_cell, field_name_cell]):
            continue
        
        # Update current address if present
        if addr_cell:
            current_address = parse_address(str(addr_cell))
        
        # Update current register if present
        if reg_name_cell:
            current_register = str(reg_name_cell).strip()
            
            # Create register if it doesn't exist
            if current_register not in registers:
                registers[current_register] = {
                    'name': current_register,
                    'addressOffset': current_address,
                    'size': 32,  # Default to 32-bit
                    'fields': []
                }
        
        # Parse field
        if field_name_cell and current_register:
            field_name = str(field_name_cell).strip()
            bit_offset_str = str(bit_offset_cell) if bit_offset_cell else '0'
            access = str(access_cell).strip().upper() if access_cell else 'RW'
            reset_value = parse_reset_value(str(reset_cell)) if reset_cell else 0
            description = str(desc_cell).strip() if desc_cell else ''
            
            # Parse bit range
            msb, lsb = parse_bit_range(bit_offset_str)
            bit_width = msb - lsb + 1
            
            field = {
                'name': field_name,
                'bitOffset': lsb,
                'bitWidth': bit_width,
                'access': access,
                'resetValue': reset_value,
                'description': description
            }
            
            registers[current_register]['fields'].append(field)
    
    return list(registers.values())


def parse_excel(file_bytes: bytes) -> Dict[str, Any]:
    """
    Main entry point for parsing Excel file.
    
    Args:
        file_bytes: Excel file content as bytes
        
    Returns:
        Dictionary containing parsed register data in IP-XACT compatible format
    """
    try:
        # Load workbook from bytes
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        
        result = {
            'version': {},
            'addressBlocks': []
        }
        
        # Parse Version sheet
        if 'Version' in wb.sheetnames:
            result['version'] = parse_version_sheet(wb['Version'])
        
        # Parse Address Map sheet
        address_blocks = []
        if 'Address Map' in wb.sheetnames:
            address_blocks = parse_address_map_sheet(wb['Address Map'])
        
        # Parse each block sheet
        for block in address_blocks:
            block_name = block['name']
            if block_name in wb.sheetnames:
                registers = parse_block_sheet(wb[block_name], block_name)
                block['registers'] = registers
            else:
                block['registers'] = []
        
        result['addressBlocks'] = address_blocks
        
        return result
        
    except Exception as e:
        raise Exception(f"Failed to parse Excel file: {str(e)}")


# For testing in Node.js/Pyodide environment
def parse_excel_json(file_bytes: bytes) -> str:
    """
    Parse Excel and return JSON string.
    This is the function that will be called from JavaScript.
    """
    result = parse_excel(file_bytes)
    return json.dumps(result)
